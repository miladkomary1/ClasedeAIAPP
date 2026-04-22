"""
run_takeoff.py  —  DXF Rebar Quantity Take-Off
===============================================
Self-contained. No relative imports. Works from any directory.

CLI:
    python run_takeoff.py <drawing.dxf> [output_dir]

Programmatic:
    from run_takeoff import run_takeoff
    out_path = run_takeoff("/path/to/drawing.dxf", "/path/to/outputs/")

Installs required:
    pip install ezdxf openpyxl --break-system-packages -q
"""

import re
import os
import sys
from collections import Counter, defaultdict
from pathlib import Path


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 1 — REBAR PATTERN MATCHING
# ══════════════════════════════════════════════════════════════════════════════
#
# All patterns use the same named groups so the rest of the pipeline is uniform:
#   qty   — number of bars  (e.g. 35)
#   dia   — diameter in mm  (e.g. 16)
#   spc   — spacing value   (e.g. 20)
#   spc_sep — spacing separator used, for faithful display (c/, @, e=, s=)
#   len   — length as written (e.g. 720  or  230-250  or  400/500)
#   mark  — position code   (e.g. R.S., R.I., RS, RI, SUP, INF — or empty)
#
# Ordered STRICT → PERMISSIVE.  First match wins.
# The Ø character class [ØΦφ∅Oo] covers all common CAD font substitutions.
# ──────────────────────────────────────────────────────────────────────────────

_DIA   = r'[ØΦφ∅Oo]'              # diameter symbol variants
_LEN   = r'\d+(?:[/\-]\d+)?'      # length: single, range (230-250), slash (400/500)
_MARK  = r'R\.[SILTCF]\.|R[SILF]|SUP|INF|EXT|INT|TOP|BOT|[Tt]op|[Bb]ot'

# Boundary after the last field — prevents matching mid-sentence dimension labels.
# Allows: end-of-string, newline, or a space followed by non-alphanumeric or end.
_END   = r'(?=\s*$|\s*[\r\n]|\s+[^A-Za-z0-9]|$)'

REBAR_PATTERNS = [

    # ── 1. Spanish canonical, MARK REQUIRED  ─────────────────────────────────
    # 35Ø16c/20 L=720 R.S.
    re.compile(
        r'(?P<qty>\d+)\s*' + _DIA + r'\s*(?P<dia>\d+)'
        r'\s*(?P<spc_sep>[cC]/)\s*(?P<spc>\d+)'
        r'\s+[Ll]=\s*(?P<len>' + _LEN + r')'
        r'\s+(?P<mark>' + _MARK + r')',
        re.IGNORECASE
    ),

    # ── 2. Spanish canonical, no mark  ───────────────────────────────────────
    # 35Ø16c/20 L=720
    # Boundary guard prevents false positives on annotations with trailing text.
    re.compile(
        r'(?P<qty>\d+)\s*' + _DIA + r'\s*(?P<dia>\d+)'
        r'\s*(?P<spc_sep>[cC]/)\s*(?P<spc>\d+)'
        r'\s+[Ll]=\s*(?P<len>' + _LEN + r')'
        r'(?P<mark>)' + _END,
        re.IGNORECASE
    ),

    # ── 3. @ spacing, mark optional  ─────────────────────────────────────────
    # 35Ø16@20 L=720 R.S.   or   35Ø16@20 L=720
    re.compile(
        r'(?P<qty>\d+)\s*' + _DIA + r'\s*(?P<dia>\d+)'
        r'\s*(?P<spc_sep>@)\s*(?P<spc>\d+)'
        r'\s+[Ll]=\s*(?P<len>' + _LEN + r')'
        r'(?:\s+(?P<mark>' + _MARK + r'))?',
        re.IGNORECASE
    ),

    # ── 4. e= / s= spacing (French/Italian/UK)  ──────────────────────────────
    # 35Ø16 e=20 L=720   or   35Ø16 s=20 L=600 R.S.
    re.compile(
        r'(?P<qty>\d+)\s*' + _DIA + r'\s*(?P<dia>\d+)'
        r'\s*(?P<spc_sep>[eEsS]=)\s*(?P<spc>\d+)'
        r'\s+[Ll]=\s*(?P<len>' + _LEN + r')'
        r'(?:\s+(?P<mark>' + _MARK + r'))?',
        re.IGNORECASE
    ),

    # ── 5. Long.= / l.= length label  ────────────────────────────────────────
    # 35Ø16c/20 Long.=720   or   35Ø16c/20 l.=720
    re.compile(
        r'(?P<qty>\d+)\s*' + _DIA + r'\s*(?P<dia>\d+)'
        r'\s*(?P<spc_sep>[cC]/)\s*(?P<spc>\d+)'
        r'\s+[Ll](?:ong)?\.=\s*(?P<len>' + _LEN + r')'
        r'(?:\s+(?P<mark>' + _MARK + r'))?',
        re.IGNORECASE
    ),

    # ── 6. L: colon separator  ────────────────────────────────────────────────
    # 35Ø16c/20 L:720 RS
    re.compile(
        r'(?P<qty>\d+)\s*' + _DIA + r'\s*(?P<dia>\d+)'
        r'\s*(?P<spc_sep>[cC]/)\s*(?P<spc>\d+)'
        r'\s+[Ll]:\s*(?P<len>' + _LEN + r')'
        r'(?:\s+(?P<mark>' + _MARK + r'))?',
        re.IGNORECASE
    ),

]

# MTEXT formatting-code stripper  (e.g. \fArial|b1|i0|c0|p34|;)
# Note: remove { and } individually — do NOT remove the content between them,
# since the rebar annotation text is often inside the outermost braces.
_MTEXT_RE = re.compile(r'\\[a-zA-Z][^;]*;|\\\S|[{}]')


def clean_mtext(raw: str) -> str:
    """Remove AutoCAD MTEXT inline formatting codes."""
    return _MTEXT_RE.sub('', raw).strip()


def parse_rebar_note(text: str):
    """
    Try all patterns against `text`.
    Returns dict with keys: qty, dia, spc, spc_sep, len_raw, mark
    or None if no pattern matched.
    """
    cleaned = clean_mtext(text)
    for pat in REBAR_PATTERNS:
        m = pat.search(cleaned)
        if m:
            gd = m.groupdict()
            return {
                'qty':     int(gd['qty']),
                'dia':     int(gd['dia']),
                'spc':     gd['spc'],
                'spc_sep': gd.get('spc_sep', 'c/') or 'c/',
                'len_raw': gd['len'],
                'mark':    (gd.get('mark') or '').strip(),
            }
    return None


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 2 — DXF TEXT EXTRACTION  (INSERT-aware)
# ══════════════════════════════════════════════════════════════════════════════
#
# CRITICAL: iterating doc.blocks gives each text entity once per DEFINITION.
# If a detail block is INSERT-ed 5 times in modelspace, texts inside it should
# be counted 5 times.  We walk the modelspace tree following INSERT entities.
# ──────────────────────────────────────────────────────────────────────────────

def _collect_texts_from_block(doc, block_name, depth=0, max_depth=25):
    """
    Yield raw text strings from all TEXT/MTEXT entities inside `block_name`,
    recursively following nested INSERT entities.
    `depth` guard prevents infinite loops on malformed/circular DXF files.
    """
    if depth > max_depth:
        return
    if block_name not in doc.blocks:
        return
    for entity in doc.blocks[block_name]:
        etype = entity.dxftype()
        if etype == 'MTEXT':
            try:
                t = entity.text
                if t and t.strip():
                    yield t.strip()
            except Exception:
                pass
        elif etype == 'TEXT':
            try:
                t = entity.dxf.text
                if t and t.strip():
                    yield t.strip()
            except Exception:
                pass
        elif etype == 'INSERT':
            try:
                sub = entity.dxf.name
                yield from _collect_texts_from_block(doc, sub, depth + 1, max_depth)
            except Exception:
                pass


def extract_all_texts(dxf_path: str):
    """
    Return list of ALL text strings in the drawing, with correct INSERT counts.
    Texts inside a block that is inserted N times appear N times in the result.
    """
    try:
        import ezdxf
    except ImportError:
        raise ImportError(
            "ezdxf not installed. Run: pip install ezdxf --break-system-packages"
        )
    doc = ezdxf.readfile(dxf_path)
    texts = []

    # Walk modelspace (and paper space layouts) top-level entities
    for layout in doc.layouts:
        for entity in layout:
            etype = entity.dxftype()
            if etype == 'MTEXT':
                try:
                    t = entity.text
                    if t and t.strip():
                        texts.append(t.strip())
                except Exception:
                    pass
            elif etype == 'TEXT':
                try:
                    t = entity.dxf.text
                    if t and t.strip():
                        texts.append(t.strip())
                except Exception:
                    pass
            elif etype == 'INSERT':
                try:
                    block_name = entity.dxf.name
                    texts.extend(
                        _collect_texts_from_block(doc, block_name, depth=0)
                    )
                except Exception:
                    pass

    return texts


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 3 — LENGTH CONVERSION
# ══════════════════════════════════════════════════════════════════════════════

def to_meters(length_raw: str):
    """
    Parse the raw length string and convert to metres.
    Returns (length_m: float, is_range: bool, assumed_unit: str).

    Unit auto-detection heuristic:
      - val > 2000  → assume mm  (divide by 1000)
      - val ≤ 2000  → assume cm  (divide by 100)
    Range (e.g. "230-250" or "400/500"): average of the two values.
    """
    raw = str(length_raw).strip()

    # Detect range separators
    for sep in ('-', '/'):
        if sep in raw:
            parts = raw.split(sep, 1)
            try:
                a, b = float(parts[0]), float(parts[1])
                val = (a + b) / 2
                unit = 'mm' if val > 2000 else 'cm'
                divisor = 1000 if unit == 'mm' else 100
                return round(val / divisor, 4), True, unit
            except ValueError:
                pass  # fall through to single-value parse

    # Single value
    try:
        val = float(raw)
    except ValueError:
        return 0.0, False, '?'

    unit = 'mm' if val > 2000 else 'cm'
    divisor = 1000 if unit == 'mm' else 100
    return round(val / divisor, 4), False, unit


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 4 — EXCEL BUILDER
# ══════════════════════════════════════════════════════════════════════════════

def build_rebar_xlsx(rows: list, output_path: str, drawing_name: str = ''):
    """
    Write a two-sheet formatted Excel workbook.
    Sheet 1: full rebar schedule, one row per unique annotation.
    Sheet 2: totals grouped by diameter.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ImportError(
            "openpyxl not installed. Run: pip install openpyxl --break-system-packages"
        )

    # ── Palette ──────────────────────────────────────────────────────────────
    C_DARK  = '1F3864'
    C_MID   = '2F5496'
    C_ALT   = 'EBF3FB'
    C_WHITE = 'FFFFFF'
    C_AMBER = 'E8A202'
    C_FLAG  = 'FFF3CD'   # pale yellow for range-length rows
    C_NOTE  = 'F7F7F7'   # light grey for subtitle

    def _f(bold=False, size=9, color='000000', italic=False):
        return Font(name='Arial', bold=bold, size=size,
                    color=color, italic=italic)

    def _fill(hex_col):
        return PatternFill('solid', fgColor=hex_col)

    _thin  = Side(style='thin',   color='CCCCCC')
    _thick = Side(style='medium', color=C_MID)
    BDR    = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    BDR_TK = Border(left=_thick, right=_thick, top=_thick, bottom=_thick)

    AL_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
    AL_L = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    COLS = list('ABCDEFGHIJKLMN')  # 14 columns

    def cell(ws, col_idx, row, value=None, bold=False, size=9, fg='000000',
             italic=False, fill=None, align=None, border=BDR, fmt=None):
        c = ws[f'{COLS[col_idx]}{row}']
        if value is not None:
            c.value = value
        c.font = _f(bold=bold, size=size, color=fg, italic=italic)
        if fill:
            c.fill = _fill(fill)
        c.alignment = align or AL_C
        if border:
            c.border = border
        if fmt:
            c.number_format = fmt
        return c

    wb = Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1 — Rebar Schedule
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = 'Rebar Schedule'

    # Title
    ws.merge_cells('A1:N1')
    cell(ws, 0, 1,
         value=f'REBAR QUANTITY TAKE-OFF  ·  {drawing_name}',
         bold=True, size=12, fg=C_WHITE, fill=C_DARK, align=AL_C, border=None)
    ws.row_dimensions[1].height = 26

    # Subtitle
    ws.merge_cells('A2:N2')
    cell(ws, 0, 2,
         value='Bar lengths assumed in cm unless >2000 (then mm)  ·  '
               'Unit weight = Ø²÷162  ·  🟡 yellow = range length (average used)  ·  '
               'R.I.=bottom  R.S.=top  ·  Occurrences = INSERT-aware count',
         size=8, italic=True, fg='555555', fill=C_NOTE, align=AL_L, border=None)
    ws.row_dimensions[2].height = 14

    # Column headers
    headers = [
        ('Rebar Note',         22, AL_L),
        ('Occurrences\n(×)',    9, AL_C),
        ('Qty/\nNote',          8, AL_C),
        ('Total\nBars',         8, AL_C),
        ('Ø\n(mm)',             7, AL_C),
        ('Spacing',             9, AL_C),
        ('L/bar\n(cm)',         9, AL_C),
        ('L/bar\n(m)',          9, AL_C),
        ('Total L\n(m)',       11, AL_C),
        ('Unit W\n(kg/m)',     11, AL_C),
        ('Total W\n(kg)',      11, AL_C),
        ('Total W\n(t)',       10, AL_C),
        ('Mark',                7, AL_C),
        ('Flags',              14, AL_L),
    ]
    ws.row_dimensions[3].height = 34
    for ci, (hdr, width, al) in enumerate(headers):
        cell(ws, ci, 3, value=hdr, bold=True, size=9, fg=C_WHITE,
             fill=C_MID, align=al)
        ws.column_dimensions[COLS[ci]].width = width

    # Data rows
    for ri, row in enumerate(rows):
        r = ri + 4
        flags = row.get('flags', [])
        row_fill = C_FLAG if row['is_range_length'] else (C_ALT if ri % 2 == 0 else C_WHITE)

        row_data = [
            (row['note'],             AL_L, None),
            (row['occurrences'],      AL_C, None),
            (row['qty_per_note'],     AL_C, None),
            (row['total_bars'],       AL_C, None),
            (row['dia_mm'],           AL_C, None),
            (row['spacing_display'],  AL_C, None),
            (row['length_per_bar_raw'], AL_C, None),
            (row['length_per_bar_m'], AL_C, '0.00'),
            (row['total_length_m'],   AL_C, '#,##0.00'),
            (row['unit_weight'],      AL_C, '0.0000'),
            (row['total_weight_kg'],  AL_C, '#,##0.00'),
            (row['total_weight_t'],   AL_C, '0.0000'),
            (row['mark'],             AL_C, None),
            ('; '.join(flags),        AL_L, None),
        ]
        for ci, (val, al, fmt) in enumerate(row_data):
            cell(ws, ci, r, value=val, fill=row_fill, align=al, fmt=fmt)
        ws.row_dimensions[r].height = 15

    # Totals row
    tr = len(rows) + 4
    ws.merge_cells(f'A{tr}:D{tr}')
    n_notes = len(rows)
    n_bars  = sum(r['total_bars'] for r in rows)
    cell(ws, 0, tr,
         value=f'TOTALS  ·  {n_notes} unique note types  ·  {n_bars} total bars',
         bold=True, size=10, fg=C_WHITE, fill=C_DARK, align=AL_L, border=BDR_TK)
    for ci in range(4, 14):
        cell(ws, ci, tr, fill=C_DARK, fg=C_WHITE, border=BDR_TK)

    total_len = round(sum(r['total_length_m'] for r in rows), 2)
    total_kg  = round(sum(r['total_weight_kg'] for r in rows), 2)
    total_t   = round(total_kg / 1000, 4)

    cell(ws, 8,  tr, value=total_len, bold=True, size=10, fg=C_WHITE,
         fill=C_DARK, border=BDR_TK, fmt='#,##0.00')
    cell(ws, 10, tr, value=total_kg,  bold=True, size=10, fg=C_WHITE,
         fill=C_AMBER, border=BDR_TK, fmt='#,##0.00')
    cell(ws, 11, tr, value=total_t,   bold=True, size=10, fg=C_WHITE,
         fill=C_AMBER, border=BDR_TK, fmt='0.000')
    ws.row_dimensions[tr].height = 20

    ws.freeze_panes = 'A4'

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2 — By Diameter
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('By Diameter')
    ws2.merge_cells('A1:G1')
    ws2['A1'].value       = 'REBAR SUMMARY BY DIAMETER'
    ws2['A1'].font        = _f(bold=True, size=12, color=C_WHITE)
    ws2['A1'].fill        = _fill(C_DARK)
    ws2['A1'].alignment   = AL_C
    ws2.row_dimensions[1].height = 24

    hdrs2 = ['Ø (mm)', 'Note Types', 'Total Bars',
             'Total Length (m)', 'Unit Weight (kg/m)',
             'Total Weight (kg)', 'Total Weight (t)']
    widths2 = [10, 12, 12, 18, 18, 18, 14]
    for ci, (h, w) in enumerate(zip(hdrs2, widths2)):
        c2 = ws2[f'{COLS[ci]}2']
        c2.value     = h
        c2.font      = _f(bold=True, size=9, color=C_WHITE)
        c2.fill      = _fill(C_MID)
        c2.alignment = AL_C
        c2.border    = BDR
        ws2.column_dimensions[COLS[ci]].width = w
    ws2.row_dimensions[2].height = 28

    by_dia = defaultdict(lambda: {'notes': 0, 'bars': 0,
                                   'length': 0.0, 'weight': 0.0})
    for row in rows:
        d = row['dia_mm']
        by_dia[d]['notes']  += 1
        by_dia[d]['bars']   += row['total_bars']
        by_dia[d]['length'] += row['total_length_m']
        by_dia[d]['weight'] += row['total_weight_kg']

    for ri2, (dia, data) in enumerate(sorted(by_dia.items(), reverse=True)):
        r2 = ri2 + 3
        uw   = round(dia ** 2 / 162, 4)
        fill2 = C_ALT if ri2 % 2 == 0 else C_WHITE
        vals2  = [dia, data['notes'], data['bars'],
                  round(data['length'], 2), uw,
                  round(data['weight'], 2),
                  round(data['weight'] / 1000, 4)]
        fmts2  = [None, None, None, '#,##0.00',
                  '0.0000', '#,##0.00', '0.0000']
        for ci, (v, f) in enumerate(zip(vals2, fmts2)):
            c2 = ws2[f'{COLS[ci]}{r2}']
            c2.value     = v
            c2.font      = _f(size=9)
            c2.fill      = _fill(fill2)
            c2.alignment = AL_C
            c2.border    = BDR
            if f:
                c2.number_format = f
        ws2.row_dimensions[r2].height = 15

    # Grand total row
    tr2 = len(by_dia) + 3
    ws2.merge_cells(f'A{tr2}:C{tr2}')
    for ci in range(7):
        ws2[f'{COLS[ci]}{tr2}'].fill   = _fill(C_DARK)
        ws2[f'{COLS[ci]}{tr2}'].border = BDR_TK
        ws2[f'{COLS[ci]}{tr2}'].font   = _f(bold=True, size=10, color=C_WHITE)
        ws2[f'{COLS[ci]}{tr2}'].alignment = AL_C
    ws2[f'A{tr2}'].value = 'GRAND TOTAL'

    grand_len = round(sum(d['length'] for d in by_dia.values()), 2)
    grand_kg  = round(sum(d['weight'] for d in by_dia.values()), 2)
    ws2[f'D{tr2}'].value = grand_len
    ws2[f'D{tr2}'].number_format = '#,##0.00'
    ws2[f'F{tr2}'].value = grand_kg
    ws2[f'F{tr2}'].number_format = '#,##0.00'
    ws2[f'F{tr2}'].fill   = _fill(C_AMBER)
    ws2[f'G{tr2}'].value = round(grand_kg / 1000, 4)
    ws2[f'G{tr2}'].number_format = '0.000'
    ws2[f'G{tr2}'].fill   = _fill(C_AMBER)
    ws2.row_dimensions[tr2].height = 20

    wb.save(output_path)


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 5 — MAIN PIPELINE
# ══════════════════════════════════════════════════════════════════════════════

def run_takeoff(dxf_path: str, output_dir: str = None) -> str:
    """
    Full pipeline: read DXF → extract texts → match patterns → compute → write XLSX.
    Returns the path to the output .xlsx file.
    """
    dxf_path = str(dxf_path)
    stem = Path(dxf_path).stem

    print(f'\n{"="*68}')
    print(f'DXF Rebar Take-Off  ·  {Path(dxf_path).name}')
    print(f'{"="*68}')

    # ── Step 1: Extract texts ─────────────────────────────────────────────────
    print('Reading DXF and collecting text entities...')
    texts = extract_all_texts(dxf_path)
    print(f'  Total text entities: {len(texts)}')

    # ── Step 2: Match rebar notes ─────────────────────────────────────────────
    note_counter = Counter()
    for t in texts:
        if parse_rebar_note(t):
            note_counter[t.strip()] += 1

    if not note_counter:
        print('\n⚠  No rebar annotations matched.')
        print('   Check references/patterns.md for all supported notation styles.')
        print('   Tip: print a sample of the raw texts with extract_all_texts().')
        return None, []

    print(f'  Unique rebar annotations matched: {len(note_counter)}')

    # ── Step 3: Build rows ────────────────────────────────────────────────────
    rows = []
    for note, occurrences in sorted(note_counter.items()):
        p = parse_rebar_note(note)
        if not p:
            continue

        bar_m, is_range, assumed_unit = to_meters(p['len_raw'])
        total_bars = occurrences * p['qty']
        total_m    = round(total_bars * bar_m, 2)
        uw         = round(p['dia'] ** 2 / 162, 4)
        total_kg   = round(total_m * uw, 2)

        # Spacing display — preserve original notation
        spc_sep = p['spc_sep'] if p['spc_sep'] else 'c/'
        spc_display = f"{spc_sep}{p['spc']}"

        # Build flags list for the Flags column
        flags = []
        if is_range:
            flags.append(f'range avg ({p["len_raw"]})')
        if assumed_unit == 'mm':
            flags.append('length in mm (auto-detected)')

        rows.append({
            'note':               note,
            'occurrences':        occurrences,
            'qty_per_note':       p['qty'],
            'total_bars':         total_bars,
            'dia_mm':             p['dia'],
            'spacing_display':    spc_display,
            'length_per_bar_raw': p['len_raw'],
            'length_per_bar_m':   bar_m,
            'total_length_m':     total_m,
            'unit_weight':        uw,
            'total_weight_kg':    total_kg,
            'total_weight_t':     round(total_kg / 1000, 4),
            'mark':               p['mark'],
            'is_range_length':    is_range,
            'flags':              flags,
        })

    # Sort: diameter descending, then note alphabetically
    rows.sort(key=lambda r: (-r['dia_mm'], r['note']))

    # ── Step 4: Console summary ───────────────────────────────────────────────
    print(f'\n{"─"*68}')
    print(f'  {"Annotation":<40} {"Bars":>5}  {"Length(m)":>10}  {"Weight(kg)":>10}')
    print(f'{"─"*68}')
    for row in rows:
        flag = '  ⚠' if row['flags'] else ''
        print(f'  {row["note"]:<40} {row["total_bars"]:>5}  '
              f'{row["total_length_m"]:>10.2f}  {row["total_weight_kg"]:>10.2f}{flag}')

    total_bars_all = sum(r['total_bars'] for r in rows)
    total_m_all    = sum(r['total_length_m'] for r in rows)
    total_kg_all   = sum(r['total_weight_kg'] for r in rows)
    print(f'{"─"*68}')
    print(f'  {"TOTAL":<40} {total_bars_all:>5}  '
          f'{total_m_all:>10.2f}  {total_kg_all:>10.2f}  '
          f'({total_kg_all/1000:.3f} t)')

    flags_rows = [r for r in rows if r['flags']]
    if flags_rows:
        print(f'\n  ⚠  {len(flags_rows)} row(s) flagged — check Flags column in Excel.')

    # ── Step 5: Write XLSX ────────────────────────────────────────────────────
    if output_dir is None:
        output_dir = os.path.dirname(dxf_path) or '.'
    os.makedirs(output_dir, exist_ok=True)
    out_path = os.path.join(output_dir, f'rebar_takeoff_{stem}.xlsx')

    build_rebar_xlsx(rows, out_path, drawing_name=stem)
    print(f'\n  Saved: {out_path}')
    print(f'{"="*68}\n')

    return out_path, rows


# ── CLI entry point ────────────────────────────────────────────────────────────
if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('Usage: python run_takeoff.py <drawing.dxf> [output_dir]')
        sys.exit(1)
    run_takeoff(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
